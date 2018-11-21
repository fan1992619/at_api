#18210542401 bearer     Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NTIyNCwiZXhwIjoxNTY1NDIxMjI0LCJuYmYiOjE1MzM4ODUyMjQsImp0aSI6Im80ZmFKVnhlMUxXSmlmdksiLCJzdWIiOjMxLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.UCbB_ilaw1xu92aB4oQ5k_dJXT6tj-xHVGtO5-NiHTM
#18782610762 bearer     Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NDY0NywiZXhwIjoxNTY1NDIwNjQ3LCJuYmYiOjE1MzM4ODQ2NDcsImp0aSI6ImlGTlJkTTRSVk90S3ZhWTgiLCJzdWIiOjMzLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.uMkAI8VR6lZOCr27znRYLfkZRazvpvxDHjc8wBi1xPw
from lxml import html
# import xlsxwriter
# import urllib.request
from openpyxl import load_workbook,Workbook
import requests
import urllib.request
import random
import xlrd
from openpyxl.drawing.image import Image
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
from data import data_config_article
import json
import time
header_at= {
    'Host': 'api.at.top',
    'Connection': 'Keep-Alive',
    'Accept-Encoding': 'gzip',
    'User-Agent': 'okhttp/3.8.1',
    'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTU0MjcxMDM2MywiZXhwIjoxNTc0MjQ2MzYzLCJuYmYiOjE1NDI3MTAzNjMsImp0aSI6Ijk4ZUtCb2gzaEV1SUE1ckgiLCJzdWIiOiIzMSIsInBydiI6ImM4ZWUxZmM4OWU3NzVlYzRjNzM4NjY3ZTViZTE3YTU5MGI2ZDQwZmMifQ.Ey2Ot4nRgH_fV8Q7D42aKoXH2NzzPYja6bedpBqaXI4',
    'deviceid': 'ac:c1:ee:c0:33:34-ac:c1:ee:c0:33:34',
    'getuiclientid': '5b9a0d6f110d2b136f9ca135d93fad06',
    'platform': 'android',
    'userid': '33',
    'version': '2.1.0'
}
num = random.randint(1, 18)
str_182='Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NTIyNCwiZXhwIjoxNTY1NDIxMjI0LCJuYmYiOjE1MzM4ODUyMjQsImp0aSI6Im80ZmFKVnhlMUxXSmlmdksiLCJzdWIiOjMxLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.UCbB_ilaw1xu92aB4oQ5k_dJXT6tj-xHVGtO5-NiHTM'
str_187='Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NDY0NywiZXhwIjoxNTY1NDIwNjQ3LCJuYmYiOjE1MzM4ODQ2NDcsImp0aSI6ImlGTlJkTTRSVk90S3ZhWTgiLCJzdWIiOjMzLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.uMkAI8VR6lZOCr27znRYLfkZRazvpvxDHjc8wBi1xPw'
class SendArticles():
    def spider_article(self):
        wb = load_workbook('../dataconfig/title.xlsx')
        ws = wb.active
        ws.title = "抓取标题"
        ws.sheet_properties.tabColor = 'ff0000'
        url = 'https://www.jinse.com/news/bitcoin'
        res = requests.get(url=url, verify=False)
        res.encoding = 'utf-8'
        html_doc = res.text
        # 获取xpath对象
        selector = html.fromstring(html_doc)
        # 找到列表集合
        ui_list = selector.xpath(
            '//div[@class="wrap margin-b clearfix"]/div[@class="wrap-left left"]/div/div[@class="content"]/div[@class="list clearfix"]')
        for i, li in enumerate(ui_list):
            title = li.xpath('//div[@class="post right"]/div/a/@title')
            link = li.xpath('//div[@class="post right"]/div/a/@href')
            content=li.xpath('//div[@class="post right"]/div[@class="message"]/text()')
            img_url=li.xpath('//div[@class="image left"]/a/img/@src')
            # print(content[i])
            # print('第%d次' % i,title[i],link[i],content[i],img_url[i])
            ws['A{0}'.format(i + 2)] = title[i]
            ws['B{0}'.format(i + 2)] = link[i]
            ws['C{0}'.format(i + 2)]=content[i]
            ws['D{0}'.format(i + 2)]=img_url[i]
            i += 1
        wb.save('../dataconfig/title.xlsx')
    #获取url
    def open_url(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../dataconfig/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config_article.get_url())
        url =tables.cell_value(row, col)
        # print(url)
        return url
    #获取title
    def open_title(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../dataconfig/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config_article.get_title())
        title =tables.cell_value(row, col)
        # print(title)
        return title
    #获取文章内容
    def open_content(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../dataconfig/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config_article.get_content())
        content =tables.cell_value(row, col)
        return content
    #获取文章图片的url
    def open_img_url(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../dataconfig/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config_article.get_img_url())
        img_url =tables.cell_value(row, col)
        # print(title)
        return img_url
    #获取这张图片内容
    def open_img_content(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../dataconfig/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config_article.get_img_content())
        img_content =tables.cell_value(row, col)
        # print(title)
        return img_content
    #遍历url
    def articles_url(self):
        self.spider_article()
        data = xlrd.open_workbook('../dataconfig/title.xlsx')
        tables_num = data.sheets()[0].nrows
        for i in range(tables_num):
            if i<1:
                continue
            return self.open_url(i)
    #发布项目文章
    def send_article_project(self):
        num = random.randint(1, 18)
        #随机获取项目id
        articles=[22,23,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,1503]
        random_articles_num=random.choice(articles)
        self.spider_article()
        link=self.open_url(num)
        title=self.open_title(num)
        #发布文章到项目
        url_test_project='https://api.at.top/v1/projects/{0}/articles'.format(random_articles_num)
        data={
            'type': '1',
            'link': '{0}'.format(link),
            'title':'{0}'.format(title)
            # 'timestamp': '1537583500',
            # 'signature': '43fa84601a94e0b5920b5ce7a73c680f'
        }
        res= requests.post(url=url_test_project, data=data, headers=header_at, verify=False).json()
        time.sleep(3)
        # print("发布项目文章",res)
        res=json.dumps(res)
        return res
    #发布主题文章
    def send_article_subject(self):
        num = random.randint(1, 18)
        # 随机获取主题id
        subjects = [1, 2, 3, 4, 5, 6]
        random_subject = random.choice(subjects)
        self.spider_article()
        link = self.open_url(num)
        title = self.open_title(num)
        url_test_subject = 'https://api.at.top/v1/subjects/{0}/articles'.format(random_subject)
        data = {
            'type': '1',
            'link': '{0}'.format(link),
            'title': '{0}'.format(title)
            # 'timestamp': '1537583500',
            # 'signature': '43fa84601a94e0b5920b5ce7a73c680f'
        }
        res = requests.post(url=url_test_subject, data=data, headers=header_at, verify=False).json()
        time.sleep(3)
        # print("发布主题文章",res)
        res = json.dumps(res)
        return res
    #发布提问到项目
    def send_question_project(self):
        num = random.randint(1, 18)
        self.spider_article()
        #随机获取项目id
        project_id=[22,23,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,1503]
        random_project_num=random.choice(project_id)
        content = self.open_content(num)
        title = self.open_title(num)+"?"
        url_question_project = 'https://api.at.top/v1/question'
        data_project={
            'project_id': random_project_num,
            'is_anonymous': 1,
            'description': '{0}'.format(content),
            'title': '{0}'.format(title)
        }
        res = requests.post(url=url_question_project, data=data_project, headers=header_at, verify=False).json()
        # print("发布提问到项目",res)
        res = json.dumps(res)
        return res
    #发布提问到主题
    def send_question_subject(self):
        num = random.randint(1, 18)
        self.spider_article()
        #随机获取主题id
        subjects = [1, 2, 3, 4, 5, 6]
        random_subject=random.choice(subjects)
        content = self.open_content(num)
        title = self.open_title(num)+"?"
        url_question_subject = 'https://api.at.top/v1/question/subject'
        data_subject = {
            'subject_id':random_subject,
            'is_anonymous':1,
            'description': '{0}'.format(content),
            'title': '{0}'.format(title)
        }
        res = requests.post(url=url_question_subject, data=data_subject, headers=header_at, verify=False).json()
        # print("发布提问到主题",res)
        res = json.dumps(res)
        return res
    #回答api发布评论
    def api_commit_deep(self):
        num = random.randint(1, 10)
        self.spider_article()
        comment_content=self.open_content(num)
        #获取评论的内容
        comment_data={
            'content':comment_content
        }
        time.sleep(5)
        url_comment='https://api.at.top/v1/comment/answer/45717909825137196'
        res=requests.post(url=url_comment, headers=header_at, data=comment_data).json()
        if res['code']==-204 or res['code']==-2:
            url_comment = 'https://api.at.top/v1/comment/answer/42810485585750572'
            res=requests.post(url=url_comment, headers=header_at, data=comment_data).json()
        else:
            pass
        return json.dumps(res)
if __name__ == '__main__':
    send = SendArticles()
    # print(send.open_url(3))
    # print(send.open_title(3))
    # send.spider_images()
    # print(send.articles_url())
    # print(send.open_img_url(3))
    # print(send.open_content(3))
    # send.send_article_subject()
    # send.send_atricle_project()
    send.send_question_project()
    send.send_question_subject()
    # send.api_commit_deep()
    # send.send_get()